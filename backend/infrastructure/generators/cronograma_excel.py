"""
cronograma_excel.py — Diseño final v4 (layout corregido)
==========================================
Layout (fila 0-based drawingML):

  ROW 0           : Título "Cronograma del Proyecto" + subtítulo
  ROW 1..N_PILLS  : Pills de roles (wrap automático)
  ROW N_PILLS+1   : Meses (col B+) | Card Info (col A, span múltiples filas)
  ROW N_PILLS+2   : Semanas
  ROW N_PILLS+3   : Sprints
  ROW N_PILLS+4+  : Actividades
"""

import io
import math
import zipfile
import datetime
from lxml import etree
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ── Negocio ───────────────────────────────────────────────────────────────────
HORAS_SEMANALES_DEFAULT = 42
SEMANAS_POR_MES    = 4
SEMANAS_POR_SPRINT = 2

# ── Paleta ────────────────────────────────────────────────────────────────────
VERDES_PRIMARIOS = ["166534", "15803D", "16A34A", "22C55E", "059669", "047857"]
BARRA_COLORES = VERDES_PRIMARIOS
ROLE_COLORS   = ["15803D", "16A34A", "047857", "166534", "22C55E", "059669"]

COLOR_GESTION   = "047857"
COLOR_PREP      = "166534"
COLOR_MES       = "15803D"
COLOR_SEMANA_BG = "DCFCE7"
COLOR_SEMANA_TX = "166534"
COLOR_SPRINT    = "059669"
COLOR_KO        = "14532D"
COLOR_INFO_BG   = "F0FDF4"
COLOR_INFO_BOR  = "16A34A"
COLOR_TITULO    = "14532D"
COLOR_SUBTITULO = "166534"

# ── EMU columnas ──────────────────────────────────────────────────────────────
PADDING     =   38_100
COL_A_EMU   = 1_600_200   # col A – etiquetas actividades
COL_B_EMU   =   457_200   # col B – kick off / S0 / Sprint 0
COL_SEM_EMU =   457_200   # cada semana

# ── EMU filas ─────────────────────────────────────────────────────────────────
ROW_TITULO_EMU  =   571_500
ROW_PILL_EMU    =   342_900
ROW_HDR_EMU     =   342_900
ROW_SUB_EMU     =   285_750
ROW_SPRINT_EMU  =   285_750
ROW_ACT_EMU     =   314_325


# ── API pública ───────────────────────────────────────────────────────────────
def generate_cronograma(config: dict) -> bytes:
    actividades  = config.get("actividades", [])
    roles_raw    = config.get("roles", [])
    nombre_proy  = config.get("nombre_proyecto", "Proyecto")
    torre_proy   = config.get("torre", "")
    id_proy      = config.get("id_proyecto", "")
    fecha_str    = config.get("fecha", datetime.date.today().strftime("%d de %B de %Y"))
    horas_semanales = _horas_semanales(config.get("horas_semanales"))

    if not actividades:
        raise ValueError("No hay actividades")

    roles = _normalizar_roles(roles_raw)

    total_horas = 0
    for act in actividades:
        p = max(1, int(act.get("personas", 1)))
        act["semanas"] = max(1, math.ceil(act["horas"] / horas_semanales / p))
        total_horas += act["horas"]

    total_semanas  = max(a["semanas"] for a in actividades)
    duracion_meses = round(total_semanas / SEMANAS_POR_MES, 1)
    pill_extra_cols = _pill_extra_columns(roles, total_semanas)

    n_roles     = len(roles)
    filas_pills = 1 if n_roles else 0  # siempre una sola fila de pills

    # ROW 0 = título
    # ROW 1 .. filas_pills = pills
    # ROW filas_pills+1    = meses / kick-off  ← ROW_HDR_START
    ROW_HDR_START = 1 + filas_pills

    meta = {
        "nombre_proyecto": nombre_proy,
        "torre":           torre_proy,
        "id_proyecto":     id_proy,
        "fecha":           fecha_str,
        "total_horas":     total_horas,
        "duracion_meses":  duracion_meses,
        "horas_semanales": horas_semanales,
        "filas_pills":     filas_pills,
        "ROW_HDR_START":   ROW_HDR_START,
    }

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Cronograma"
    ws1.sheet_view.showGridLines = False
    _configurar_dimensiones(ws1, total_semanas, filas_pills,
                            sin_semanas=False, extra_cols=pill_extra_cols)

    ws2 = wb.create_sheet(title="Cronograma sin semanas")
    ws2.sheet_view.showGridLines = False
    _configurar_dimensiones(ws2, total_semanas, filas_pills,
                            sin_semanas=True, extra_cols=pill_extra_cols)

    buf = io.BytesIO()
    wb.save(buf)

    return _inyectar_drawing(buf.getvalue(), actividades, roles,
                             total_semanas, meta)


def _horas_semanales(value):
    try:
        horas = float(value)
    except (TypeError, ValueError):
        return HORAS_SEMANALES_DEFAULT
    return horas if horas > 0 else HORAS_SEMANALES_DEFAULT


def _normalizar_roles(roles_raw):
    roles = []
    for r in roles_raw:
        if isinstance(r, dict):
            nombre = str(r.get("perfil", r.get("rol", ""))).strip()
            if nombre:
                roles.append({
                    "perfil":    nombre,
                    "seniority": str(r.get("seniority", "")).strip(),
                    "personas":  max(1, int(r.get("personas", 1))),
                })
        else:
            nombre = str(r).strip()
            if nombre:
                roles.append({"perfil": nombre, "seniority": "", "personas": 1})
    return roles


# ── Dimensiones ───────────────────────────────────────────────────────────────
def _configurar_dimensiones(ws, total_semanas, filas_pills, sin_semanas,
                            extra_cols=0):
    r = 1
    ws.row_dimensions[r].height = 44   # título
    r += 1
    for _ in range(filas_pills):
        ws.row_dimensions[r].height = 26
        r += 1
    ws.row_dimensions[r].height = 26;   r += 1   # meses / kick-off
    if not sin_semanas:
        ws.row_dimensions[r].height = 22; r += 1  # semanas
    ws.row_dimensions[r].height = 22;   r += 1   # sprints
    for rr in range(r, r + 60):
        ws.row_dimensions[rr].height = 24

    ws.column_dimensions["A"].width = 23
    ws.column_dimensions["B"].width = 6
    for i in range(total_semanas + extra_cols):
        ws.column_dimensions[get_column_letter(i + 3)].width = 6


def _pill_widths(roles):
    """Calcula el ancho de cada pill sin recortar su etiqueta."""
    widths = []
    for rol in roles:
        label = (f"  {rol['perfil']} {rol['seniority']}  "
                 if rol["seniority"] else f"  {rol['perfil']}  ")
        widths.append(max(300_000, len(label) * 85_000))
    return widths


def _pill_extra_columns(roles, total_semanas):
    """Reserva columnas adicionales cuando los perfiles exceden el ancho base."""
    if not roles:
        return 0
    base_w = COL_A_EMU + COL_B_EMU + total_semanas * COL_SEM_EMU
    required_w = PADDING * 4 + sum(_pill_widths(roles)) + 76_200 * (len(roles) - 1)
    return max(0, math.ceil((required_w - base_w) / COL_SEM_EMU))


# ── Inyección ZIP ─────────────────────────────────────────────────────────────
def _inyectar_drawing(xlsx_bytes, actividades, roles, total_semanas, meta):
    d1 = _build_drawing(actividades, roles, total_semanas, meta, sin_semanas=False)
    d2 = _build_drawing(actividades, roles, total_semanas, meta, sin_semanas=True)

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zin:
        files = {n: zin.read(n) for n in zin.namelist()}

    files["xl/drawings/drawing1.xml"] = d1.encode("utf-8")
    files["xl/drawings/drawing2.xml"] = d2.encode("utf-8")
    _patch_sheet(files, "sheet1.xml", "drawing1.xml")
    _patch_sheet(files, "sheet2.xml", "drawing2.xml")
    _patch_ct(files, "drawing1.xml")
    _patch_ct(files, "drawing2.xml")

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    return out.getvalue()


def _patch_sheet(files, sheet_name, drawing_name):
    rid = "rId10"
    rel_path = f"xl/worksheets/_rels/{sheet_name}.rels"
    rel_entry = (
        f'<Relationship Id="{rid}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
        f'Target="../drawings/{drawing_name}"/>'
    )
    if rel_path not in files:
        files[rel_path] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            + rel_entry + '</Relationships>'
        ).encode("utf-8")
    else:
        rel = files[rel_path].decode("utf-8")
        if drawing_name not in rel:
            rel = rel.replace("</Relationships>", rel_entry + "</Relationships>")
            files[rel_path] = rel.encode("utf-8")

    sp = f"xl/worksheets/{sheet_name}"
    sheet = files[sp].decode("utf-8")
    if "<drawing " not in sheet:
        if "xmlns:r=" not in sheet:
            sheet = sheet.replace("<worksheet ",
                '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ', 1)
        sheet = sheet.replace("</worksheet>", f'<drawing r:id="{rid}"/></worksheet>')
        files[sp] = sheet.encode("utf-8")


def _patch_ct(files, drawing_name):
    ct = files["[Content_Types].xml"].decode("utf-8")
    ov = (f'<Override PartName="/xl/drawings/{drawing_name}" '
          f'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
    if ov not in ct:
        ct = ct.replace("</Types>", ov + "</Types>")
        files["[Content_Types].xml"] = ct.encode("utf-8")


# ── Drawing principal ─────────────────────────────────────────────────────────
def _build_drawing(actividades, roles, total_semanas, meta, sin_semanas):
    XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    A   = "http://schemas.openxmlformats.org/drawingml/2006/main"
    root = etree.Element(f"{{{XDR}}}wsDr",
                         nsmap={"xdr": XDR, "a": A,
                                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"})
    sid = 1

    ROW_HDR_START = meta["ROW_HDR_START"]

    # ── Índices de filas (0-based) ─────────────────────────────────────────
    ROW_TITULO = 0
    # Pills: filas 1 .. filas_pills
    ROW_MES    = ROW_HDR_START          # meses + kick-off
    if sin_semanas:
        ROW_SEMANA = None
        ROW_SPRINT = ROW_MES + 1
    else:
        ROW_SEMANA = ROW_MES + 1
        ROW_SPRINT = ROW_MES + 2
    n_hdr = ROW_SPRINT + 1              # primera fila de actividades

    # Columnas (0-based)
    COL_KO = 1    # col B = kick-off / S0 / Sprint 0
    COL_S1 = 2    # col C en adelante = semanas

    # Ancho total: se amplía para que los perfiles permanezcan en una fila.
    base_w = COL_A_EMU + COL_B_EMU + total_semanas * COL_SEM_EMU
    pill_widths = _pill_widths(roles)
    pill_required_w = PADDING * 4 + sum(pill_widths) + 76_200 * (len(roles) - 1)
    TOTAL_W = max(base_w, pill_required_w)

    # ── 1. TÍTULO ──────────────────────────────────────────────────────────
    subtitulo = (f"Generado el {meta['fecha']}  |  "
                 f"Total: {meta['total_horas']}h  |  "
                 f"Duración: {meta['duracion_meses']} meses")

    sid = _shape_abs(root, sid,
                     "Cronograma del Proyecto",
                     row=ROW_TITULO, x=PADDING*2, y=PADDING*2,
                     w=TOTAL_W - PADDING*4, h=ROW_TITULO_EMU - PADDING*4,
                     color="FFFFFF", text_color=COLOR_TITULO,
                     font_sz=1400, bold=True, geom="rect",
                     subtitulo=subtitulo, sub_sz=900, sub_color=COLOR_SUBTITULO)

    # ── 2. PILLS DE ROLES ─────────────────────────────────────────────────
    PILL_H       = 228_600
    PILL_GAP_H   = 76_200
    PILL_GAP_V   = (ROW_PILL_EMU - PILL_H) // 2   # centrado vertical en la fila
    BADGE_R      = 133_350
    PILL_PAD_X   = PADDING * 2

    for rol, pill_w in zip(roles, pill_widths):
        label = (f"  {rol['perfil']} {rol['seniority']}  "
                 if rol["seniority"]
                 else f"  {rol['perfil']}  ")
        rol["_pill_w"] = pill_w
        rol["_label"]  = label.strip()

    x_cursor = PILL_PAD_X
    for i, rol in enumerate(roles):
        pw    = rol["_pill_w"]
        color = ROLE_COLORS[i % len(ROLE_COLORS)]

        sid = _shape_abs(root, sid, rol["_label"],
                         row=1, x=x_cursor, y=PILL_GAP_V,
                         w=pw, h=PILL_H,
                         color=color, text_color="FFFFFF",
                         font_sz=780, bold=True, geom="roundRect", adj="50000")

        # Badge número
        bx = x_cursor + pw - BADGE_R // 2
        by = PILL_GAP_V - BADGE_R // 2
        sid = _shape_abs(root, sid, str(rol["personas"]),
                         row=1, x=bx, y=by,
                         w=BADGE_R, h=BADGE_R,
                         color="10B981", text_color="FFFFFF",
                         font_sz=620, bold=True, geom="ellipse")

        x_cursor += pw + PILL_GAP_H

    # ── 3. CARD DE INFO DEL PROYECTO (col A, span filas MES→SPRINT) ───────
    # Calcular altura total del card: meses + semanas (opcional) + sprints
    if sin_semanas:
        info_h = ROW_HDR_EMU + ROW_SPRINT_EMU
    else:
        info_h = ROW_HDR_EMU + ROW_SUB_EMU + ROW_SPRINT_EMU

    info_h -= PADDING * 2   # pequeño margen

    # Fondo del card
    sid = _shape_abs(root, sid, "",
                     row=ROW_MES, x=PADDING, y=PADDING,
                     w=COL_A_EMU - PADDING*2, h=info_h,
                     color=COLOR_INFO_BG, text_color=COLOR_INFO_BG,
                     geom="roundRect", adj="17948",
                     borde=COLOR_INFO_BOR)

    # Textos del card (todos dentro de ROW_MES con offsets Y)
    _txt_y = PADDING * 3

    # "INFORMACIÓN DEL PROYECTO"
    sid = _shape_abs(root, sid, "INFORMACIÓN DEL PROYECTO",
                     row=ROW_MES, x=PADDING*3, y=_txt_y,
                     w=COL_A_EMU - PADDING*5, h=ROW_HDR_EMU // 2,
                     color=COLOR_INFO_BG, text_color="8B5CF6",
                     font_sz=640, bold=False, geom="rect")
    _txt_y += ROW_HDR_EMU // 2 + PADDING

    # Nombre del proyecto
    sid = _shape_abs(root, sid, meta["nombre_proyecto"],
                     row=ROW_MES, x=PADDING*3, y=_txt_y,
                     w=COL_A_EMU - PADDING*5, h=ROW_HDR_EMU,
                     color=COLOR_INFO_BG, text_color=COLOR_TITULO,
                     font_sz=1000, bold=True, geom="rect")
    _txt_y += ROW_HDR_EMU + PADDING

    # Torre
    if meta["torre"]:
        sid = _shape_abs(root, sid, f"□ {meta['torre']}",
                         row=ROW_MES, x=PADDING*3, y=_txt_y,
                         w=COL_A_EMU - PADDING*5, h=ROW_SUB_EMU,
                         color=COLOR_INFO_BG, text_color="8B5CF6",
                         font_sz=750, bold=False, geom="rect")
        _txt_y += ROW_SUB_EMU + PADDING

    # ID
    if meta["id_proyecto"]:
        sid = _shape_abs(root, sid, f"ID: {meta['id_proyecto']}",
                         row=ROW_MES, x=PADDING*3, y=_txt_y,
                         w=COL_A_EMU - PADDING*5, h=ROW_SUB_EMU,
                         color=COLOR_INFO_BG, text_color=COLOR_SUBTITULO,
                         font_sz=720, bold=False, geom="rect")

    # ── 4. KICK OFF (col B, fila MES) ──────────────────────────────────────
    sid = _shape(root, sid, "Kick Off",
                 COL_KO, ROW_MES, COL_KO, ROW_MES,
                 COLOR_KO, "FFFFFF", bold=True, font_sz=900,
                 col_w=COL_B_EMU, row_h=ROW_HDR_EMU)

    if ROW_SEMANA is not None:
        sid = _shape(root, sid, "S0",
                     COL_KO, ROW_SEMANA, COL_KO, ROW_SEMANA,
                     COLOR_SEMANA_BG, COLOR_SEMANA_TX, bold=False, font_sz=800,
                     col_w=COL_B_EMU, row_h=ROW_SUB_EMU)

    sid = _shape(root, sid, "Sprint 0",
                 COL_KO, ROW_SPRINT, COL_KO, ROW_SPRINT,
                 COLOR_SPRINT, "FFFFFF", bold=True, font_sz=800,
                 col_w=COL_B_EMU, row_h=ROW_SPRINT_EMU)

    # ── 5. MESES ───────────────────────────────────────────────────────────
    col_cur = COL_S1
    for m in range(math.ceil(total_semanas / SEMANAS_POR_MES)):
        col_end = min(col_cur + SEMANAS_POR_MES - 1, COL_S1 + total_semanas - 1)
        sid = _shape(root, sid, f"Mes {m+1}",
                     col_cur, ROW_MES, col_end, ROW_MES,
                     COLOR_MES, "FFFFFF", bold=True, font_sz=950,
                     row_h=ROW_HDR_EMU)
        col_cur += SEMANAS_POR_MES

    # ── 6. SEMANAS ─────────────────────────────────────────────────────────
    if ROW_SEMANA is not None:
        for s in range(total_semanas):
            c = COL_S1 + s
            sid = _shape(root, sid, f"S{s+1}",
                         c, ROW_SEMANA, c, ROW_SEMANA,
                         COLOR_SEMANA_BG, COLOR_SEMANA_TX, bold=False, font_sz=800,
                         row_h=ROW_SUB_EMU)

    # ── 7. SPRINTS ─────────────────────────────────────────────────────────
    sprint_n = 1
    for s in range(0, total_semanas, SEMANAS_POR_SPRINT):
        cs = COL_S1 + s
        ce = min(cs + SEMANAS_POR_SPRINT - 1, COL_S1 + total_semanas - 1)
        sid = _shape(root, sid, f"Sprint {sprint_n}",
                     cs, ROW_SPRINT, ce, ROW_SPRINT,
                     COLOR_SPRINT, "FFFFFF", bold=True, font_sz=850,
                     row_h=ROW_SPRINT_EMU)
        sprint_n += 1

    # ── 8. PREPARAR AMBIENTES ──────────────────────────────────────────────
    prep_row = n_hdr
    sid = _shape(root, sid, "Preparar Ambientes",
                 0, prep_row, 0, prep_row,
                 COLOR_PREP, "FFFFFF", bold=True, font_sz=900,
                 col_w=COL_A_EMU, row_h=ROW_ACT_EMU, adj="17948")
    # Barra de prep en kick-off (col B)
    sid = _shape(root, sid, "",
                 COL_KO, prep_row, COL_KO, prep_row,
                 COLOR_PREP, "FFFFFF",
                 col_w=COL_B_EMU, row_h=ROW_ACT_EMU, adj="50000")

    # ── 9. ACTIVIDADES ─────────────────────────────────────────────────────
    for i, act in enumerate(actividades):
        row   = n_hdr + 1 + i
        color = BARRA_COLORES[(i + 1) % len(BARRA_COLORES)]

        # Etiqueta (col A)
        sid = _shape(root, sid, act["torre"],
                     0, row, 0, row,
                     color, "FFFFFF", bold=True, font_sz=900,
                     col_w=COL_A_EMU, row_h=ROW_ACT_EMU, adj="17948")

        # Barra (empieza en COL_S1, no en COL_KO)
        bar_end = COL_S1 + act["semanas"] - 1
        sid = _shape(root, sid, "",
                     COL_S1, row, bar_end, row,
                     color, "FFFFFF",
                     row_h=ROW_ACT_EMU, adj="50000")

    # ── 10. GESTIÓN DEL PROYECTO ───────────────────────────────────────────
    gestion_row = n_hdr + 1 + len(actividades)
    total_fin   = COL_S1 + total_semanas - 1
    sid = _shape(root, sid, "Gestión del Proyecto",
                 0, gestion_row, 0, gestion_row,
                 COLOR_GESTION, "FFFFFF", bold=True, font_sz=900,
                 col_w=COL_A_EMU, row_h=ROW_ACT_EMU, adj="17948")
    # Barra de gestión: desde kick-off hasta fin
    sid = _shape(root, sid, "",
                 COL_KO, gestion_row, total_fin, gestion_row,
                 COLOR_GESTION, "FFFFFF",
                 row_h=ROW_ACT_EMU, adj="50000")

    return etree.tostring(root, xml_declaration=True,
                          encoding="UTF-8", standalone=True).decode("utf-8")


# ── _shape: alineada a celdas (twoCellAnchor) ────────────────────────────────
def _shape(parent, sid, text,
           col_from, row_from, col_to, row_to,
           color, text_color="FFFFFF",
           bold=False, font_sz=900,
           col_w=COL_SEM_EMU, row_h=ROW_ACT_EMU,
           adj="default"):
    XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    anchor = etree.SubElement(parent, f"{{{XDR}}}twoCellAnchor", editAs="oneCell")

    frm = etree.SubElement(anchor, f"{{{XDR}}}from")
    etree.SubElement(frm, f"{{{XDR}}}col").text    = str(col_from)
    etree.SubElement(frm, f"{{{XDR}}}colOff").text = str(PADDING)
    etree.SubElement(frm, f"{{{XDR}}}row").text    = str(row_from)
    etree.SubElement(frm, f"{{{XDR}}}rowOff").text = str(PADDING)

    to = etree.SubElement(anchor, f"{{{XDR}}}to")
    etree.SubElement(to, f"{{{XDR}}}col").text    = str(col_to)
    etree.SubElement(to, f"{{{XDR}}}colOff").text = str(col_w - PADDING)
    etree.SubElement(to, f"{{{XDR}}}row").text    = str(row_to)
    etree.SubElement(to, f"{{{XDR}}}rowOff").text = str(row_h - PADDING)

    return _fill_sp(anchor, sid, text, color, text_color, font_sz, bold, adj)


# ── _shape_abs: coordenadas absolutas dentro de una fila ─────────────────────
def _shape_abs(parent, sid, text,
               row, x, y, w, h,
               color, text_color="FFFFFF",
               font_sz=800, bold=True,
               geom="roundRect", adj="default",
               subtitulo=None, sub_sz=800, sub_color="6B7280",
               borde=None):
    XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    anchor = etree.SubElement(parent, f"{{{XDR}}}twoCellAnchor", editAs="oneCell")

    col_f, xoff_f = _x_to_col(int(x))
    col_t, xoff_t = _x_to_col(int(x + w))

    frm = etree.SubElement(anchor, f"{{{XDR}}}from")
    etree.SubElement(frm, f"{{{XDR}}}col").text    = str(col_f)
    etree.SubElement(frm, f"{{{XDR}}}colOff").text = str(xoff_f)
    etree.SubElement(frm, f"{{{XDR}}}row").text    = str(row)
    etree.SubElement(frm, f"{{{XDR}}}rowOff").text = str(int(y))

    to = etree.SubElement(anchor, f"{{{XDR}}}to")
    etree.SubElement(to, f"{{{XDR}}}col").text    = str(col_t)
    etree.SubElement(to, f"{{{XDR}}}colOff").text = str(xoff_t)
    etree.SubElement(to, f"{{{XDR}}}row").text    = str(row)
    etree.SubElement(to, f"{{{XDR}}}rowOff").text = str(int(y + h))

    return _fill_sp(anchor, sid, text, color, text_color, font_sz, bold, adj, geom,
                    subtitulo=subtitulo, sub_sz=sub_sz, sub_color=sub_color, borde=borde)


def _x_to_col(x_emu: int):
    """Convierte EMU absoluto (desde el borde izquierdo) a (col, offset)."""
    if x_emu <= 0:
        return 0, 0
    if x_emu <= COL_A_EMU:
        return 0, int(x_emu)
    x = x_emu - COL_A_EMU
    if x <= COL_B_EMU:
        return 1, int(x)
    x -= COL_B_EMU
    col = 2 + x // COL_SEM_EMU
    off = x % COL_SEM_EMU
    return int(col), int(off)


# ── _fill_sp ──────────────────────────────────────────────────────────────────
def _fill_sp(anchor, sid, text, color, text_color,
             font_sz, bold, adj="default", geom="roundRect",
             subtitulo=None, sub_sz=800, sub_color="6B7280",
             borde=None):
    XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    A   = "http://schemas.openxmlformats.org/drawingml/2006/main"

    sp     = etree.SubElement(anchor, f"{{{XDR}}}sp", macro="", textlink="")
    nvSpPr = etree.SubElement(sp, f"{{{XDR}}}nvSpPr")
    etree.SubElement(nvSpPr, f"{{{XDR}}}cNvPr", id=str(sid), name=f"s{sid}")
    etree.SubElement(nvSpPr, f"{{{XDR}}}cNvSpPr")

    spPr     = etree.SubElement(sp, f"{{{XDR}}}spPr")
    prstGeom = etree.SubElement(spPr, f"{{{A}}}prstGeom", prst=geom)
    avLst    = etree.SubElement(prstGeom, f"{{{A}}}avLst")
    if adj != "default":
        etree.SubElement(avLst, f"{{{A}}}gd", name="adj", fmla=f"val {adj}")

    sf = etree.SubElement(spPr, f"{{{A}}}solidFill")
    etree.SubElement(sf, f"{{{A}}}srgbClr", val=color)

    ln = etree.SubElement(spPr, f"{{{A}}}ln", w="12700")
    if borde:
        sf_ln = etree.SubElement(ln, f"{{{A}}}solidFill")
        etree.SubElement(sf_ln, f"{{{A}}}srgbClr", val=borde)
    else:
        etree.SubElement(ln, f"{{{A}}}noFill")

    txBody = etree.SubElement(sp, f"{{{XDR}}}txBody")
    anchor_val = "t" if subtitulo else "ctr"
    etree.SubElement(txBody, f"{{{A}}}bodyPr",
                     wrap="square", rtlCol="0", anchor=anchor_val,
                     lIns="91440", rIns="91440", tIns="45720", bIns="45720")
    etree.SubElement(txBody, f"{{{A}}}lstStyle")

    if text:
        p = etree.SubElement(txBody, f"{{{A}}}p")
        etree.SubElement(p, f"{{{A}}}pPr", algn="l" if subtitulo else "ctr")
        r     = etree.SubElement(p, f"{{{A}}}r")
        attrs = {"lang": "es-CO", "sz": str(font_sz), "dirty": "0"}
        if bold:
            attrs["b"] = "1"
        rPr = etree.SubElement(r, f"{{{A}}}rPr", **attrs)
        sf2 = etree.SubElement(rPr, f"{{{A}}}solidFill")
        etree.SubElement(sf2, f"{{{A}}}srgbClr", val=text_color)
        etree.SubElement(rPr, f"{{{A}}}latin", typeface="Calibri")
        etree.SubElement(r, f"{{{A}}}t").text = text

    if subtitulo:
        p2 = etree.SubElement(txBody, f"{{{A}}}p")
        etree.SubElement(p2, f"{{{A}}}pPr", algn="l", spcBef="100000")
        r2   = etree.SubElement(p2, f"{{{A}}}r")
        rPr2 = etree.SubElement(r2, f"{{{A}}}rPr",
                                 lang="es-CO", sz=str(sub_sz), dirty="0")
        sf3  = etree.SubElement(rPr2, f"{{{A}}}solidFill")
        etree.SubElement(sf3, f"{{{A}}}srgbClr", val=sub_color)
        etree.SubElement(rPr2, f"{{{A}}}latin", typeface="Calibri")
        etree.SubElement(r2, f"{{{A}}}t").text = subtitulo

    if not text and not subtitulo:
        p = etree.SubElement(txBody, f"{{{A}}}p")
        etree.SubElement(p, f"{{{A}}}pPr", algn="ctr")

    etree.SubElement(anchor, f"{{{XDR}}}clientData")
    return sid + 1
